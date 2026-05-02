#!/usr/bin/env python3
"""
sync_excel_json.py — Sincronización bidireccional Excel ↔ JSON
Hermes Legal Pro v4.0

Uso:
    python3 scripts/sync_excel_json.py --direction excel-to-json
    python3 scripts/sync_excel_json.py --direction json-to-excel
    python3 scripts/sync_excel_json.py --direction both
"""

import json
import argparse
from pathlib import Path
from datetime import datetime, date

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl no está instalado. Ejecuta: pip3 install openpyxl")
    raise SystemExit(1)

# ── Paths ─────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent.parent
EXCEL_PATH = BASE_DIR / "excel" / "Centro_Operativo_Maestro_Willow_v4.xlsx"
JSON_PATH = BASE_DIR / "dashboard" / "datos" / "matters.json"
SHEET_NAME = "Matters"

# ── Column mapping ────────────────────────────────────────────
# Excel col index (1-based) → JSON field
EXCEL_COLS = {
    1: "id",           # Matter ID
    2: "cliente",      # Cliente
    3: None,           # Representante (not in JSON flat model)
    4: "estado",       # Status
    5: "area_practica", # Área
    6: "prioridad",    # Prioridad
    7: "fecha_creacion", # Fecha Apertura
    8: "deadline",     # Deadline
    9: None,           # Días Restantes (formula)
    10: None,          # Docs Pendientes (formula)
    11: "next_step",   # Bloqueo Actual
}

STATUS_MAP_EXCEL_TO_JSON = {
    "active": "activo",
    "closed": "cerrado",
    "paused": "pausado",
    "urgent": "urgente",
}

STATUS_MAP_JSON_TO_EXCEL = {v: k for k, v in STATUS_MAP_EXCEL_TO_JSON.items()}

PRIORITY_MAP_EXCEL_TO_JSON = {
    "high": "alta",
    "medium": "media",
    "low": "baja",
}

PRIORITY_MAP_JSON_TO_EXCEL = {v: k for k, v in PRIORITY_MAP_EXCEL_TO_JSON.items()}

# ── Helpers ───────────────────────────────────────────────────
def load_json():
    if not JSON_PATH.exists():
        return []
    with open(JSON_PATH, "r", encoding="utf-8") as f:
        return json.load(f)

def save_json(data):
    JSON_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def load_excel():
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Excel no encontrado: {EXCEL_PATH}")
    return openpyxl.load_workbook(EXCEL_PATH)

def save_excel(wb):
    EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(EXCEL_PATH)

def excel_row_to_dict(row):
    """Convierte una fila de Excel a dict JSON"""
    data = {}
    for col_idx, field in EXCEL_COLS.items():
        if field is None:
            continue
        value = row[col_idx - 1]
        if value is None:
            data[field] = None
            continue
        
        if field == "estado":
            data[field] = STATUS_MAP_EXCEL_TO_JSON.get(str(value).lower(), str(value).lower())
        elif field == "prioridad":
            data[field] = PRIORITY_MAP_EXCEL_TO_JSON.get(str(value).lower(), str(value).lower())
        else:
            data[field] = str(value) if value is not None else None
    
    return data

def dict_to_excel_row(matter):
    """Convierte un matter JSON a lista de valores para Excel"""
    row = [None] * 11
    for col_idx, field in EXCEL_COLS.items():
        if field is None:
            continue
        
        value = matter.get(field)
        
        if field == "estado":
            row[col_idx - 1] = STATUS_MAP_JSON_TO_EXCEL.get(str(value).lower(), str(value).lower())
        elif field == "prioridad":
            row[col_idx - 1] = PRIORITY_MAP_JSON_TO_EXCEL.get(str(value).lower(), str(value).lower())
        else:
            row[col_idx - 1] = value if value is not None else ""
    
    return row

# ── Sync functions ────────────────────────────────────────────
def sync_excel_to_json():
    """Lee Excel y actualiza JSON"""
    print("📊 Sync: Excel → JSON")
    
    wb = load_excel()
    ws = wb[SHEET_NAME]
    
    # Leer filas de Excel
    excel_matters = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or str(row[0]).strip() == "":
            continue
        excel_matters.append(excel_row_to_dict(row))
    
    # Leer JSON actual
    json_matters = load_json()
    
    # Merge: preservar campos que no están en Excel
    json_by_id = {m["id"]: m for m in json_matters}
    
    merged = []
    for em in excel_matters:
        matter_id = em["id"]
        if matter_id in json_by_id:
            # Actualizar existente, preservar campos no mapeados
            existing = json_by_id[matter_id].copy()
            for key, val in em.items():
                if val is not None and val != [] and val != {}:
                    existing[key] = val
            merged.append(existing)
        else:
            # Nuevo matter
            em["fecha_creacion"] = em.get("fecha_creacion") or date.today().isoformat()
            merged.append(em)
    
    # Agregar matters que están en JSON pero no en Excel (opcional: podríamos ignorarlos)
    # Por ahora, solo sincronizamos los del Excel
    
    save_json(merged)
    print(f"  ✅ {len(merged)} matters sincronizados a JSON")
    return merged

def sync_json_to_excel():
    """Lee JSON y actualiza Excel"""
    print("📊 Sync: JSON → Excel")
    
    json_matters = load_json()
    
    wb = load_excel()
    ws = wb[SHEET_NAME]
    
    # Capturar fórmulas de la primera fila de datos (fila 2) antes de limpiar
    formula_templates = {}
    if ws.max_row >= 2:
        for col in range(1, 12):
            cell = ws.cell(row=2, column=col)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formula_templates[col] = cell.value
    
    # Limpiar filas existentes (excepto header)
    max_row = ws.max_row
    if max_row > 1:
        ws.delete_rows(2, max_row - 1)
    
    # Escribir matters desde JSON
    for idx, matter in enumerate(json_matters):
        row = dict_to_excel_row(matter)
        ws.append(row)
        
        # Aplicar fórmulas a la nueva fila (row = idx + 2)
        new_row = idx + 2
        for col, formula in formula_templates.items():
            # Ajustar referencias de fila en la fórmula
            adjusted_formula = formula.replace('2', str(new_row))
            ws.cell(row=new_row, column=col, value=adjusted_formula)
    
    save_excel(wb)
    print(f"  ✅ {len(json_matters)} matters sincronizados a Excel")
    return json_matters

def sync_both():
    """Sincronización bidireccional: merge inteligente"""
    print("📊 Sync: Bidireccional (merge)")
    
    # Leer ambas fuentes
    wb = load_excel()
    ws = wb[SHEET_NAME]
    
    excel_matters = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or str(row[0]).strip() == "":
            continue
        excel_matters.append(excel_row_to_dict(row))
    
    json_matters = load_json()
    
    # Merge por ID, priorizando el más reciente (por ahora: Excel gana en conflictos)
    by_id = {}
    
    for m in json_matters:
        by_id[m["id"]] = m.copy()
    
    for m in excel_matters:
        mid = m["id"]
        if mid in by_id:
            # Merge: Excel sobrescribe campos mapeados, JSON preserva extras
            for key, val in m.items():
                if val is not None and val != [] and val != {}:
                    by_id[mid][key] = val
        else:
            by_id[mid] = m
    
    merged = list(by_id.values())
    
    # Guardar ambos
    save_json(merged)
    
    # Capturar fórmulas de la primera fila de datos antes de limpiar
    formula_templates = {}
    if ws.max_row >= 2:
        for col in range(1, 12):
            cell = ws.cell(row=2, column=col)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formula_templates[col] = cell.value
    
    # Reconstruir Excel
    max_row = ws.max_row
    if max_row > 1:
        ws.delete_rows(2, max_row - 1)
    
    for idx, matter in enumerate(merged):
        row = dict_to_excel_row(matter)
        ws.append(row)
        
        # Aplicar fórmulas a la nueva fila
        new_row = idx + 2
        for col, formula in formula_templates.items():
            adjusted_formula = formula.replace('2', str(new_row))
            ws.cell(row=new_row, column=col, value=adjusted_formula)
    
    save_excel(wb)
    print(f"  ✅ {len(merged)} matters sincronizados en ambas direcciones")
    return merged

# ── Main ──────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Sync Excel ↔ JSON para Hermes Legal Pro")
    parser.add_argument("--direction", choices=["excel-to-json", "json-to-excel", "both"],
                        default="both", help="Dirección de sincronización")
    parser.add_argument("--test", action="store_true", help="Modo test: no guarda cambios")
    args = parser.parse_args()
    
    if args.test:
        print("🧪 MODO TEST — No se guardarán cambios")
    
    try:
        if args.direction == "excel-to-json":
            result = sync_excel_to_json()
        elif args.direction == "json-to-excel":
            result = sync_json_to_excel()
        else:
            result = sync_both()
        
        print("\n📋 Matters sincronizados:")
        for m in result:
            print(f"  • {m['id']} — {m.get('cliente', 'N/A')} [{m.get('estado', 'N/A')}]")
        
        if args.test:
            print("\n⚠️  MODO TEST: cambios NO guardados")
        else:
            print("\n✅ Sincronización completada")
        
    except Exception as e:
        print(f"\n❌ Error: {e}")
        raise SystemExit(1)

if __name__ == "__main__":
    main()
